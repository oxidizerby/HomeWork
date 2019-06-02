from openpyxl import load_workbook
import json


def task7_2(xlfile, jsfile):
    wb = load_workbook(xlfile)
    sheet = wb.worksheets[0]

    students = {}
    ii = 0
    total_average_mark = 0
    for i in range(1, sheet.max_row + 1):
        sheet.cell(i, 2).value = sheet.cell(i, 1).value + ' ' + sheet.cell(i, 2).value
        marks = []
        for j in range(3, sheet.max_column + 1):
            if sheet.cell(i, j).value == 'end':
                sheet.cell(i, 1).value = f"=AVERAGE(C{i}:{sheet.cell(i, j - 1).coordinate})"
                break
            else:
                marks.append(sheet.cell(i, j).value)
        average_mark = sum(marks)/len(marks)
        students[str(sheet.cell(i, 2).value)] = {'marks': marks, 'average_mark': average_mark}
        total_average_mark += average_mark
        ii = i
    d = {'students': students, 'total_average_mark': total_average_mark/ii}

    wb.save('test_out.xlsx')
    wb.close()

    with open(jsfile, 'w', encoding="utf-8") as jsf:
        json.dump(d, jsf, indent=4, ensure_ascii=False)
        jsf.close()
    pass


xlsx_file = 'test.xlsx'
json_file = 'test.json'
task7_2(xlsx_file, json_file)
